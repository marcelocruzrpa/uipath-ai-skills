#!/usr/bin/env python3
"""
UiPath Project Scaffolder

Copies real Studio-exported templates and customizes them for a new project.
Supports three variants:
  - sequence:   Simple sequence project (non-REFramework)
  - dispatcher: REFramework Dispatcher — same base template, TransactionItem
                type changed from QueueItem to the specified type (default: DataRow).
                Claude Code must customize GetTransactionData.xaml and Process.xaml.
  - performer:  REFramework Performer (QueueItem-based) — used as-is.

Both dispatcher and performer share the same clean REFramework template
(assets/reframework/). The dispatcher variant swaps the TransactionItem type
and adds dispatcher-specific comments to Process.xaml.

The TransactionItem type for dispatchers depends on the data source:
  - DataRow:     Excel/CSV/DB data (most common)
  - String:      File paths, URLs, simple IDs
  - MailMessage: Email processing
  - Dictionary:  Complex multi-field items
  - (any type):  Whatever GetTransactionData produces

IMPORTANT: Template dependency versions are baseline defaults and go stale.
Before scaffolding, run scripts/resolve_nuget.py to get real latest versions.
Pass updated versions via --deps to override template defaults.

Usage:
    python3 scaffold_project.py --name "MyProject" --variant performer --output /path
    python3 scaffold_project.py --name "MyProject" --variant dispatcher --output /path
    python3 scaffold_project.py --name "MyProject" --variant dispatcher --transaction-type String --output /path
    python3 scaffold_project.py --name "MyProject" --variant dispatcher --transaction-type MailMessage --output /path
"""

import argparse
import json
import os
import re
import shutil
import sys
import urllib.error
import uuid
from pathlib import Path

# Ensure UTF-8 output on all platforms (Windows cmd defaults to cp1252)
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# Maps short type names to full XAML type info:
#   (xaml_type, xmlns_prefix, xmlns_uri, assembly_ref, namespace_import)
TRANSACTION_TYPE_MAP = {
    "DataRow": (
        "sd:DataRow",
        "sd",
        'clr-namespace:System.Data;assembly=System.Data.Common',
        "System.Data.Common",
        "System.Data",
    ),
    "String": (
        "x:String",
        None, None, None, None,  # x: is always available
    ),
    "MailMessage": (
        "snm:MailMessage",
        "snm",
        'clr-namespace:System.Net.Mail;assembly=System.Net.Mail',
        "System.Net.Mail",
        "System.Net.Mail",
    ),
    "Dictionary": (
        "scg:Dictionary(x:String, x:Object)",
        None, None, None, None,  # scg: already in REFramework
    ),
}


def get_skill_dir():
    """Find the skill directory (parent of scripts/)."""
    return Path(__file__).resolve().parent.parent


def _swap_transaction_type(content: str, from_type: str, to_type: str) -> str:
    """Replace TransactionItem type references in XAML content."""
    for wrapper in ['x:TypeArguments="{t}"',
                    'Type="InArgument({t})"',
                    'Type="OutArgument({t})"',
                    'Type="InOutArgument({t})"']:
        content = content.replace(
            wrapper.format(t=from_type),
            wrapper.format(t=to_type)
        )
    return content


def _ensure_xmlns(content: str, prefix: str, uri: str) -> str:
    """Add xmlns declaration if not present.

    Primary strategy: insert before the xmlns:ui declaration (stable anchor).
    Fallback: insert before the closing '>' of the <Activity element.
    """
    decl = f'xmlns:{prefix}="{uri}"'
    if decl in content:
        return content
    # Primary: insert before xmlns:ui
    anchor = 'xmlns:ui="http://schemas.uipath.com/workflow/activities"'
    if anchor in content:
        content = content.replace(anchor, f'{decl}\n  {anchor}')
    else:
        # Fallback: insert before closing > of <Activity element
        content = re.sub(
            r'(<Activity\b[^>]*?)(>)',
            rf'\1 {decl}\2',
            content, count=1,
        )
    return content


def _ensure_namespace_import(content: str, namespace: str) -> str:
    """Add namespace import to NamespacesForImplementation if not present."""
    check = f">{namespace}</x:String>"
    if namespace and check not in content:
        content = content.replace(
            '<x:String>System.Collections.Generic</x:String>',
            '<x:String>System.Collections.Generic</x:String>\n'
            f'      <x:String>{namespace}</x:String>'
        )
    return content


def _ensure_assembly_ref(content: str, assembly: str) -> str:
    """Add assembly reference to ReferencesForImplementation if not present."""
    check = f">{assembly}</AssemblyReference>"
    if assembly and check not in content:
        content = content.replace(
            '<AssemblyReference>System.Data</AssemblyReference>',
            '<AssemblyReference>System.Data</AssemblyReference>\n'
            f'      <AssemblyReference>{assembly}</AssemblyReference>'
        )
    return content


def _replace_gtd_body_for_dispatcher(gtd_path: Path, transaction_type: str,
                                      xaml_type: str):
    """Replace GetQueueItem in GetTransactionData.xaml with DataTable row indexing.

    The performer template uses RetryScope → GetQueueItem to fetch from an
    Orchestrator queue. Dispatchers iterate over a local DataTable instead,
    using in_TransactionNumber as the row index.

    For DataRow: If in_TransactionNumber <= Rows.Count → get row, else Nothing.
    For other types: Insert a placeholder comment (no GetQueueItem).
    """
    content = gtd_path.read_text(encoding="utf-8")

    # Locate the RetryScope block containing GetQueueItem
    retry_start = content.find('<ui:RetryScope DisplayName="Retry Get transaction item"')
    if retry_start == -1:
        return  # Already modified or different template

    # Find the closing tag of this RetryScope
    # Walk through nested tags to find matching close
    retry_end = content.find('</ui:RetryScope>', retry_start)
    if retry_end == -1:
        return
    retry_end = content.find('\n', retry_end)  # include the line break
    if retry_end == -1:
        retry_end = len(content)

    retry_block = content[retry_start:retry_end]

    if transaction_type == "DataRow":
        # DataTable row indexing pattern with deterministic first-call guard.
        # The If guard ensures extraction activities (inserted by replace-marker)
        # only run on the first REFramework loop iteration. Without this structural
        # guard, the LLM's extraction code would re-execute every iteration.
        replacement = (
            '<If Condition="[in_TransactionNumber = 1]" '
            'DisplayName="First transaction - load data from source" '
            'sap:VirtualizedContainerService.HintSize="494.4,400" '
            'sap2010:WorkflowViewState.IdRef="If_FirstTransaction">\r\n'
            '      <If.Then>\r\n'
            '        <Sequence DisplayName="Navigate, extract and filter" '
            'sap2010:WorkflowViewState.IdRef="Sequence_DispatcherLoad">\r\n'
            '          <sap:WorkflowViewStateService.ViewState>\r\n'
            '            <scg:Dictionary x:TypeArguments="x:String, x:Object">\r\n'
            '              <x:Boolean x:Key="IsExpanded">True</x:Boolean>\r\n'
            '            </scg:Dictionary>\r\n'
            '          </sap:WorkflowViewStateService.ViewState>\r\n'
            '          <ui:Comment DisplayName="SCAFFOLD.DISPATCHER_LOAD_DATA Load data on first call" '
            'sap2010:WorkflowViewState.IdRef="Comment_Dispatcher_1" '
            'Text="//  Replace this Comment with InvokeWorkflowFile activities:&#xA;'
            '//  1. AppName_NavigateToDataSource.xaml (navigation)&#xA;'
            '//  2. AppName_ExtractData.xaml -&gt; io_dt_TransactionData (extraction)&#xA;'
            '//  3. FilterDataTable or .Select() to filter io_dt_TransactionData" />\r\n'
            '        </Sequence>\r\n'
            '      </If.Then>\r\n'
            '    </If>\r\n'
            '    <If Condition="[in_TransactionNumber &lt;= io_dt_TransactionData.Rows.Count]" '
            'DisplayName="Check if more rows in DataTable" '
            'sap:VirtualizedContainerService.HintSize="494.4,400" '
            'sap2010:WorkflowViewState.IdRef="If_Dispatcher_1">\r\n'
            '      <If.Then>\r\n'
            '        <Assign DisplayName="Assign TransactionItem from DataTable row" '
            'sap:VirtualizedContainerService.HintSize="450,113.6" '
            'sap2010:WorkflowViewState.IdRef="Assign_Dispatcher_1">\r\n'
            '          <Assign.To>\r\n'
            f'            <OutArgument x:TypeArguments="{xaml_type}">[out_TransactionItem]</OutArgument>\r\n'
            '          </Assign.To>\r\n'
            '          <Assign.Value>\r\n'
            f'            <InArgument x:TypeArguments="{xaml_type}">[io_dt_TransactionData.Rows(in_TransactionNumber - 1)]</InArgument>\r\n'
            '          </Assign.Value>\r\n'
            '        </Assign>\r\n'
            '      </If.Then>\r\n'
            '      <If.Else>\r\n'
            '        <Assign DisplayName="Assign Nothing (end process)" '
            'sap:VirtualizedContainerService.HintSize="450,113.6" '
            'sap2010:WorkflowViewState.IdRef="Assign_Dispatcher_2">\r\n'
            '          <Assign.To>\r\n'
            f'            <OutArgument x:TypeArguments="{xaml_type}">[out_TransactionItem]</OutArgument>\r\n'
            '          </Assign.To>\r\n'
            '          <Assign.Value>\r\n'
            f'            <InArgument x:TypeArguments="{xaml_type}">[Nothing]</InArgument>\r\n'
            '          </Assign.Value>\r\n'
            '        </Assign>\r\n'
            '      </If.Else>\r\n'
            '    </If>'
        )
    else:
        # Non-DataRow dispatcher: placeholder comment
        replacement = (
            '<ui:Comment DisplayName="SCAFFOLD.DISPATCHER_GET_ITEM Retrieve transaction item" '
            'sap2010:WorkflowViewState.IdRef="Comment_Dispatcher_1" '
            f'Text="//  Replace this Comment with logic to retrieve transaction item (type: {transaction_type}).&#xA;'
            '//  Use in_TransactionNumber as the index. Set out_TransactionItem = Nothing when done." />'
        )

    content = content[:retry_start] + replacement + content[retry_end:]

    # Also remove the Config key references that are performer-only
    # (OrchestratorQueueName, OrchestratorQueueFolder, RetryNumberGetTransactionItem
    #  are no longer used in GetTransactionData for dispatchers)

    gtd_path.write_text(content, encoding="utf-8")


def _apply_dispatcher_variant(project_dir: Path, transaction_type: str):
    """
    Transform performer template into dispatcher variant.

    Changes:
    1. TransactionItem type QueueItem -> specified type in all relevant files
    2. GetTransactionData.xaml: Replace GetQueueItem with DataTable row indexing
    3. Process.xaml: Add dispatcher-specific comment
    """
    type_info = TRANSACTION_TYPE_MAP.get(transaction_type)
    if not type_info:
        raise ValueError(
            f"Unknown transaction type '{transaction_type}'. "
            f"Supported: {', '.join(TRANSACTION_TYPE_MAP.keys())}"
        )

    xaml_type, xmlns_prefix, xmlns_uri, assembly_ref, ns_import = type_info

    files_to_transform = [
        "Main.xaml",
        "Framework/GetTransactionData.xaml",
        "Framework/Process.xaml",
        # SetTransactionStatus.xaml stays as QueueItem — dispatchers pass Nothing
        "Framework/RetryCurrentTransaction.xaml",
        "Tests/GetTransactionDataTestCase.xaml",
        "Tests/ProcessTestCase.xaml",
    ]

    for rel_path in files_to_transform:
        fpath = project_dir / rel_path
        if not fpath.exists():
            continue
        content = fpath.read_text(encoding="utf-8")

        # Swap type
        content = _swap_transaction_type(content, "ui:QueueItem", xaml_type)

        # Add xmlns, namespace import, assembly ref if needed
        if xmlns_prefix and xmlns_uri:
            content = _ensure_xmlns(content, xmlns_prefix, xmlns_uri)
        if ns_import:
            content = _ensure_namespace_import(content, ns_import)
        if assembly_ref:
            content = _ensure_assembly_ref(content, assembly_ref)

        fpath.write_text(content, encoding="utf-8")

    # Fix SetTransactionStatus invokes in Main.xaml: pass Nothing (QueueItem) instead
    # of the DataRow TransactionItem. Dispatchers don't process queue items.
    main_path = project_dir / "Main.xaml"
    if main_path.exists():
        main_content = main_path.read_text(encoding="utf-8")
        sts_pattern = (
            r'(WorkflowFileName="Framework\\SetTransactionStatus\.xaml">'
            r'.*?)'
            + rf'(<InArgument x:TypeArguments="{re.escape(xaml_type)}" '
            + r'x:Key="in_TransactionItem">)\[TransactionItem\](</InArgument>)'
        )

        def _fix_sts_arg(m):
            return (
                m.group(1)
                + '<InArgument x:TypeArguments="ui:QueueItem" '
                + 'x:Key="in_TransactionItem">[Nothing]'
                + m.group(3)
            )

        main_content = re.sub(
            sts_pattern, _fix_sts_arg, main_content, flags=re.DOTALL
        )
        main_path.write_text(main_content, encoding="utf-8")

    # Replace GetTransactionData.xaml body: GetQueueItem → DataTable row indexing
    gtd_path = project_dir / "Framework" / "GetTransactionData.xaml"
    if gtd_path.exists():
        _replace_gtd_body_for_dispatcher(gtd_path, transaction_type, xaml_type)

    # Update Process.xaml comment
    process_path = project_dir / "Framework" / "Process.xaml"
    if process_path.exists():
        content = process_path.read_text(encoding="utf-8")
        content = content.replace(
            '//  Invoke steps of the process',
            f'//  DISPATCHER: TransactionItem type is {transaction_type}.&#xA;'
            '//  1. Extract fields from in_TransactionItem&#xA;'
            '//  2. Process/transform data as needed&#xA;'
            '//  3. Use AddQueueItem to push each item to Orchestrator Queue'
        )
        process_path.write_text(content, encoding="utf-8")


def scaffold_project(name: str, description: str, output_dir: str,
                     variant: str = "performer", extra_deps: dict = None,
                     attended: bool = False, expression_lang: str = "VisualBasic",
                     transaction_type: str = "DataRow",
                     queue_name: str = None, queue_folder: str = None,
                     overwrite: bool = False,
                     target: str = "both",
                     version_band: str = None):
    """Scaffold a new UiPath project from real templates."""
    skill_dir = get_skill_dir()

    # Dispatcher and performer share the same REFramework base
    variant_dir_map = {
        "sequence": "simple-sequence-template",
        "dispatcher": "reframework",
        "performer": "reframework",
    }
    template_dir = skill_dir / "assets" / variant_dir_map[variant]
    project_dir = Path(output_dir) / name

    # Detect folder duplication: if output_dir already ends with the project name,
    # the caller probably already created the folder — use output_dir directly
    if Path(output_dir).name == name:
        print(f"  [WARN] --output '{output_dir}' already ends with '{name}' -- using it directly to avoid duplicate nesting")
        project_dir = Path(output_dir)

    if not template_dir.exists():
        raise FileNotFoundError(
            f"Template not found: {template_dir}\n"
            f"  Expected template directory for variant '{variant}' at: {template_dir}\n"
            f"  Skill root: {skill_dir}\n"
            f"  Available asset dirs: {[d.name for d in (skill_dir / 'assets').iterdir() if d.is_dir()] if (skill_dir / 'assets').exists() else '(assets/ not found)'}"
        )

    if project_dir.exists():
        if not overwrite:
            raise FileExistsError(
                f"Output directory already exists: {project_dir}. "
                f"Pass --overwrite to replace it."
            )
        shutil.rmtree(project_dir)

    try:
        shutil.copytree(template_dir, project_dir)
    except PermissionError as e:
        raise PermissionError(
            f"Cannot copy template to {project_dir}: {e}\n"
            f"  Check write permissions on: {Path(output_dir)}"
        ) from e
    except OSError as e:
        raise OSError(
            f"Failed to copy template from {template_dir} to {project_dir}: {e}"
        ) from e

    # Create standard Studio directories
    (project_dir / ".entities").mkdir(exist_ok=True)
    (project_dir / ".templates").mkdir(exist_ok=True)
    tmh_dir = project_dir / ".tmh"
    tmh_dir.mkdir(exist_ok=True)
    tmh_config = tmh_dir / "config.json"
    if not tmh_config.exists():
        with open(tmh_config, "w", encoding="utf-8") as f:
            f.write('{\n  "issueKeyTestcaseValues": {}\n}')

    # Create Workflows/ folder for sub-workflows (organized by app/utility)
    workflows_dir = project_dir / "Workflows"
    workflows_dir.mkdir(exist_ok=True)

    # Copy shared utility workflows (conditional on target)
    utils_dir = workflows_dir / "Utils"
    utils_source = skill_dir / "assets" / "samples" / "common-workflows" / "Workflows" / "Utils"
    if utils_source.exists():
        BROWSER_UTILS = {"Browser_NavigateToUrl.xaml"}
        DESKTOP_UTILS = set()  # Future: desktop-only utilities
        for util_file in utils_source.glob("*.xaml"):
            name_str = util_file.name
            if target == "desktop" and name_str in BROWSER_UTILS:
                continue
            if target == "web" and name_str in DESKTOP_UTILS:
                continue
            utils_dir.mkdir(exist_ok=True)
            shutil.copy2(util_file, utils_dir / name_str)

    # Apply dispatcher-specific transformations
    if variant == "dispatcher":
        _apply_dispatcher_variant(project_dir, transaction_type)

    # --- Customize project.json ---
    pj_path = project_dir / "project.json"
    with open(pj_path, "r", encoding="utf-8") as f:
        pj = json.load(f)

    pj["name"] = name
    pj["projectId"] = str(uuid.uuid4())
    pj["description"] = description
    pj["projectVersion"] = "1.0.0"

    if extra_deps:
        pj["dependencies"].update(extra_deps)

    # Run plugin scaffold hooks (e.g. Tasks persistence support)
    from plugin_loader import load_plugins, get_scaffold_hooks
    load_plugins()
    for hook in get_scaffold_hooks():
        hook(pj)

    if version_band:
        pj["versionBand"] = version_band

    pj["runtimeOptions"]["isAttended"] = attended
    pj["expressionLanguage"] = expression_lang

    if pj.get("entryPoints"):
        pj["entryPoints"][0]["uniqueId"] = str(uuid.uuid4())
    if pj.get("designOptions", {}).get("fileInfoCollection"):
        pj["designOptions"]["fileInfoCollection"] = []

    with open(pj_path, "w", encoding="utf-8") as f:
        json.dump(pj, f, indent=2)

    # --- Customize Config.xlsx (REFramework only) ---
    config_path = project_dir / "Data" / "Config.xlsx"
    if config_path.exists():
        import openpyxl
        wb = openpyxl.load_workbook(config_path)

        # Update Settings sheet
        if "Settings" in wb.sheetnames:
            ws = wb["Settings"]
            settings_updates = {
                "logF_BusinessProcessName": name,
            }
            if queue_name:
                settings_updates["OrchestratorQueueName"] = queue_name
            if queue_folder is not None:
                settings_updates["OrchestratorQueueFolder"] = queue_folder

            for row in ws.iter_rows(min_row=2, max_col=3):
                key_cell = row[0]
                val_cell = row[1]
                if key_cell.value in settings_updates:
                    val_cell.value = settings_updates[key_cell.value]

        wb.save(config_path)
        print(f"   Config.xlsx: logF_BusinessProcessName = {name}")
        if queue_name:
            print(f"   Config.xlsx: OrchestratorQueueName = {queue_name}")
        if queue_folder is not None:
            print(f"   Config.xlsx: OrchestratorQueueFolder = {queue_folder}")

    # --- Customize Main.xaml ---
    main_path = project_dir / "Main.xaml"
    if main_path.exists():
        content = main_path.read_text(encoding="utf-8")
        if variant in ("dispatcher", "performer"):
            content = re.sub(
                r'(\[Process title\]&#xA;).*?(&#xA;&#xA;\[Process description\])',
                rf'\g<1>{name}\g<2>', content
            )
        main_path.write_text(content, encoding="utf-8")

    tx_info = f" (TransactionItem: {transaction_type})" if variant == "dispatcher" else ""
    print(f"Project scaffolded: {project_dir}")
    print(f"   Variant: {variant}{tx_info}")
    print(f"   Name: {name}")
    print(f"   Description: {description}")
    print(f"   Expression Language: {expression_lang}")
    print(f"   Attended: {attended}")
    print(f"   Dependencies: {json.dumps(pj['dependencies'], indent=4)}")
    print(f"\n   Structure:")
    for root, dirs, files in os.walk(project_dir):
        level = root.replace(str(project_dir), "").count(os.sep)
        indent = "  " * level
        print(f"  {indent}{os.path.basename(root)}/")
        for file in sorted(files):
            size = os.path.getsize(os.path.join(root, file))
            print(f"  {'  ' * (level + 1)}{file} ({size:,} bytes)")

    return str(project_dir)


def parse_deps(dep_str: str) -> dict:
    if not dep_str:
        return {}
    deps = {}
    # Split on comma or whitespace to handle both formats:
    #   "Pkg:[v],Pkg2:[v]" (comma-separated)
    #   "Pkg:[v] Pkg2:[v]" (space-separated, e.g. single quoted arg)
    for pair in re.split(r'[,\s]+', dep_str):
        pair = pair.strip()
        if ":" in pair:
            pkg, ver = pair.split(":", 1)
            v = ver.strip().strip("[]")
            deps[pkg.strip()] = f"[{v}]"
    return deps


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scaffold a UiPath project")
    parser.add_argument("--name", required=True, help="Project name")
    parser.add_argument("--description", default="UiPath automation project")
    parser.add_argument("--output", required=True, help="Output directory")
    parser.add_argument("--variant", default="sequence",
                        choices=["sequence", "dispatcher", "performer"],
                        help="sequence=simple, dispatcher=REFramework, performer=REFramework+QueueItem")
    parser.add_argument("--transaction-type", default="DataRow",
                        choices=list(TRANSACTION_TYPE_MAP.keys()),
                        help="Dispatcher TransactionItem type (default: DataRow). "
                             "Ignored for performer/sequence variants.")
    parser.add_argument("--deps", nargs="*", default=[], help="Extra deps: 'Pkg:[ver],Pkg2:[ver]' OR Pkg:[ver] Pkg2:[ver] (space-separated)")
    parser.add_argument("--attended", action="store_true")
    parser.add_argument("--lang", default="VisualBasic", choices=["VisualBasic", "CSharp"])
    parser.add_argument("--queue-name", default=None,
                        help="Orchestrator queue name (written to Config.xlsx Settings sheet). "
                             "If omitted, template default 'ProcessABCQueue' is kept.")
    parser.add_argument("--queue-folder", default=None,
                        help="Orchestrator folder for the queue (written to Config.xlsx Settings sheet).")
    parser.add_argument("--band", metavar="NN", default=None,
                        help="Target version band (e.g., 25 or 26). Stamps versionBand into "
                             "project.json and resolves baseline dependencies within the band.")
    parser.add_argument("--overwrite", action="store_true",
                        help="Delete and replace existing output directory. Without this flag, "
                             "scaffolding into an existing directory fails safely.")
    parser.add_argument("--target", default="both",
                        choices=["desktop", "web", "both"],
                        help="Target platform: desktop (skip browser utils), "
                             "web (skip desktop utils), both (default)")

    args = parser.parse_args()
    # --deps can be: ["Pkg:[v],Pkg2:[v]"] (single quoted) or ["Pkg:[v]", "Pkg2:[v]"] (space-separated)
    deps_str = ",".join(args.deps) if args.deps else ""
    extra_deps = parse_deps(deps_str)

    for pkg, ver in extra_deps.items():
        clean_ver = ver.strip("[]")
        if "-" in clean_ver:
            print(f"WARNING: {pkg} version '{clean_ver}' appears to be a prerelease "
                  f"(contains '-'). Use stable versions only.")

    # Resolve band-specific baseline deps when --band is set. --deps still wins.
    band_deps = {}
    if args.band:
        try:
            from version_band import validate_band
            validate_band(args.band)
            from resolve_nuget import resolve_packages_in_band, COMMON_PACKAGES
            print(f"Resolving baseline dependencies for band {args.band}...")
            band_deps = {pkg: f"[{ver}]"
                         for pkg, ver in resolve_packages_in_band(COMMON_PACKAGES, args.band).items()}
        except (ImportError, FileNotFoundError, json.JSONDecodeError,
                urllib.error.URLError, OSError) as e:
            print(f"WARNING: Could not resolve band {args.band} deps ({e}). "
                  f"Using template baseline.", file=sys.stderr)

    merged_deps = dict(band_deps)
    merged_deps.update(extra_deps)

    # If --band was not explicit, derive it from the resolved year-based deps.
    # This keeps project.json's versionBand in sync with what was actually
    # pinned, so lints 120-124 engage on downstream edits without the caller
    # having to pass --band. Stays silent (no stamping) when no year-based
    # dep is present so sequence scaffolds with bare deps remain opt-in.
    effective_band = args.band
    if effective_band is None and merged_deps:
        try:
            from version_band import derive_band_from_deps, UnsupportedBandError
            effective_band = derive_band_from_deps(merged_deps)
            if effective_band:
                print(f"Derived versionBand={effective_band} from resolved dependencies.")
        except UnsupportedBandError as e:
            print(f"WARNING: {e}. versionBand will be omitted.", file=sys.stderr)
        except (ImportError, ValueError) as e:
            print(f"WARNING: Could not derive versionBand from deps ({e}). "
                  f"versionBand will be omitted.", file=sys.stderr)

    try:
        scaffold_project(args.name, args.description, args.output, args.variant,
                         merged_deps if merged_deps else extra_deps,
                         args.attended, args.lang, args.transaction_type,
                         args.queue_name, args.queue_folder, args.overwrite,
                         args.target, effective_band)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    except FileExistsError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    except PermissionError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Scaffolding failed: {type(e).__name__}: {e}", file=sys.stderr)
        sys.exit(1)
